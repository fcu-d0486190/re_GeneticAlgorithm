from openpyxl import load_workbook


class ReadFile:
    def __init__(self):
        self.component = []
        self.machinename = []
        self.information = {}
        self.readfile()
    def readfile(self):
        file = load_workbook(filename=r'D:\\result\\time_model\\DMG01_TimeModel.xlsx')
        xlsx = file['Sheet']
        componentinformation = []
        componentlist = []
        for line in range(2, xlsx.max_row + 1):  # read row (line)
            for section in range(1, xlsx.max_column + 1):  # read column
                data = str(xlsx.cell(row=line, column=section).value)
                if section == 1:
                    if data not in componentlist:
                        componentlist.append(data)
                    componentinformation.append(data)
                if section == 7:
                    self.machinename.append(data)
                if section == 12:
                    worktime = data
                if section == 13:
                    changetime = data
            componentinformation.append(float(worktime)-float(changetime))
            self.component.append(componentinformation)
            componentinformation = []
        for i in componentlist:
            time = []
            for j in self.component:
                if j[0] == i:
                    time.append(j[1])
            self.information.update({i:time})

